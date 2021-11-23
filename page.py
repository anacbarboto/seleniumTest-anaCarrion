from element import EspolEducationElement
from locators import EspolEducationLocators, CareerPageLocators
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

import openpyxl
import time
import warnings

warnings.simplefilter("ignore", ResourceWarning)


class BasePage(object):
    """Base class to initialize the base page that will be called from all
    pages"""

    def __init__(self, driver, list_href, careers_name):
        self.driver = driver
        self.list_href = list_href
        self.careers_name = careers_name


class EspolEducationPage(BasePage):
    def get_faculties(self):
        return WebDriverWait(self.driver, 20).until(EC.visibility_of_all_elements_located(EspolEducationLocators.FACULTY_LIST))

    def get_classes(self, codigo):
        return self.driver.find_elements(By.XPATH, '//div[@id="'+codigo+'"]//div[@class="field-content"]//ul//li//a')

    def create_excel_file(self):

        wb = openpyxl.Workbook()
        hoja = wb.create_sheet("Hoja")
        hoja.append(('Carrera', 'Facultad', 'Link'))

        principal_window = self.driver.current_window_handle
        facultades = self.get_faculties()
        href = []
        nombres_facultades = []

        #Leer nombres de las facultades
        i = 0
        c = 0
        for facultad in facultades:
            codigo = facultad.get_attribute("href").split("#")[-1]

            i = i + 1
            if (i < 5):
                self.driver.execute_script("arguments[0].scrollIntoView(true);", facultad)
                self.driver.execute_script("arguments[0].click();", facultad)
                ActionChains(self.driver).click_and_hold(facultad).perform()
            else:
                self.driver.execute_script("window.scrollTo(0, 0);")
                self.driver.execute_script("arguments[0].click();", facultad)
                
            #Leer nombres de las materias
            if (facultad.get_attribute("aria-expanded")):
                todas_las_materias = self.get_classes(codigo)
                for materia in todas_las_materias:

                    href.append(materia.get_attribute("href"))
                    nombres_facultades.append(materia.text)

                    c = c + 1
                    hoja.cell(row=c, column=1).value = materia.text
                    hoja.cell(row=c, column=2).value= facultad.text
                    hoja.cell(row=c, column=3).value= materia.get_attribute("href")
                    print(materia.text)
                    print(facultad.text)
                    print(materia.get_attribute("href"))
    
        wb.save('educacion_espol.xlsx')
        wb.close()
        self.list_href = href
        self.careers_name = nombres_facultades
    
    def show_abet_list(self):
        abet_list = []
        list_href = self.list_href
        j = 0
        for href in list_href:
            self.driver.get(href)
            try:
                if (j != 25):
                    if (j == 31):
                        abet = CareerEducationPage.is_oceanography_abet(self)
                    else:
                        abet = CareerEducationPage.is_abet(self)
                    
                    if 'abet' in abet.text:
                        abet_list.append(self.careers_name[j])
                        print("es abet")
                else:
                    abet = CareerEducationPage.is_food_abet(self)
                    abet_list.append(self.careers_name[j])
                    print("es abet")
            except TimeoutException as ex:
                print("no es abet")
            j = j + 1
        print("Las carreras con abet son:")
        for element in abet_list:
            print(element)

class CareerEducationPage(BasePage):
    def is_abet(self):
        # Probably should search for this text in the specific page
        # element, but as for now it works fine
        return WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(CareerPageLocators.ABET_HREF))
    
    def is_oceanography_abet(self):
        return WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(CareerPageLocators.ABET_OCEANOGRAPHY))
    
    def is_food_abet(self):
        return WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(CareerPageLocators.ABET_FOOD))